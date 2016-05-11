# JSON XLSX -> JSON
# interp: a script that takes a json with Waldayu data
#         and a spreadsheet with audio filenames and
#         merges the audio filenames with the existing
#         json.

# _*_ coding:utf-8 _*_
# Import Libraries
from openpyxl import load_workbook 
from openpyxl.styles import Font, Fill, Color
from openpyxl.cell import Cell
import sys
import json
import glob
from pprint import pprint
from jsonmerge import merge

wb_audio = load_workbook(sys.argv[1])
ws = wb_audio.active
with open (sys.argv[2]) as data_file:
    data = json.load(data_file)


# Loop through all entries, find audio, check with json and insert audio into recording column
audioJson = []

y=1
for entry in ws.iter_rows('A2:D53'):
    y=y+1
    for col in entry:
        if col.column == 'A':
            audioWord = col.value
            for word in data["fm2014"]["data"]:
                if audioWord == word["word"][0]:
                    word["audio"] = ws.cell(row=y, column=4).value
                    audioJson.append(word)

# Add all data from fm2014 & git_2013 to data                    
consolidated_data =[]
for entry in data["fm2014"]["data"]:
    consolidated_data.append(entry)

for entry in data["git_2013"]["data"]:
    consolidated_data.append(entry)

# create json
json1 = json.dumps(audioJson)
json2 = json.dumps(consolidated_data)

# write both to new file, for Python 3.x change write parameter to 'w'
textfile = open('audio.json','wb')
textfile.write(json1)
textfile.close()

textfile1 = open('wordlist.json','wb')
textfile1.write(json2)
textfile1.close()

# merge json files into one, overwrite duplicates
merged_data = merge(textfile, textfile1)