# Import Libraries
from openpyxl import load_workbook 
from openpyxl.styles import Font, Fill, Color
from openpyxl.cell import Cell
import sys
import json

# Load argv[1] as workbook
wb = load_workbook(sys.argv[1])
ws = wb.active

# Create wordlist
wordList = []

# Loop through rows in worksheet, create if statements for different columns and append Lemmas to wordList.
for entry in ws.iter_rows('A2:C3'):
    newLemma = {"word":[], "definition":[]}
    for col in entry:
        if col.column == 'A':
            newLemma["word"].append(col.value)
        if col.column == 'B':
            newLemma["definition"].append(col.value)
    wordList.append(newLemma)

# create json
json = json.dumps(wordList)

# write to new file, for Python 3.x change write parameter to 'w'
textfile = open('wordlist.txt','wb')
textfile.write(json)
textfile.close()

